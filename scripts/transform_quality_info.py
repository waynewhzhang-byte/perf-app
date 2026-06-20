#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按《人员信息表字段拆分与重构设计说明书》(deepseek_markdown_20260620_d75d32.md)
修改《基本素质信息》.xlsx，输出《基本素质信息_修改》.xlsx。

规则要点：
  Sheet1「技能等级 职称」(部门 D列) -> 4.1：命中白名单则工区=该名/部门清空；否则工区=总部/部门保留。
            新列「工区」紧跟部门列(D)之后。
  Sheet2「考核结果」(所属单位 D列) -> 4.2：按是否含「/」分情况，新增「工区/部门/班组」三列，
            紧跟所属单位列(D)之后；原所属单位列永久保留。
  保持所有行数、格式、其它列不变。
"""
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

SRC = "《基本素质信息》.xlsx"
DST = "《基本素质信息_修改》.xlsx"

WHITELIST = {
    "变电检修中心", "二次检修中心", "晋北运维分部", "晋南运维分部", "晋中运维分部",
    "设备状态测试中心", "特高压北岳站", "特高压大同站", "特高压洪善站",
    "特高压雁门关换流站", "特高压长治站", "智能运检管控中心",
}


def process_sheet1_dept(dept):
    """4.1 -> (工区, 新部门)"""
    if dept is None:
        dept = ""
    s = str(dept).strip()
    if s in WHITELIST:
        return s, ""
    return "总部", s  # 空白/非白名单 -> 总部, 部门保留(原始去空白)


def process_sheet2_unit(unit):
    """4.2 -> (工区, 部门, 班组)"""
    if unit is None:
        return "总部", "", ""
    s = str(unit).strip()
    if s == "":
        return "总部", "", ""
    if "/" in s:
        left, right = s.split("/", 1)
        left, right = left.strip(), right.strip()
        if left in WHITELIST:
            return left, "", right
        return "总部", left, right
    # 不含斜杠
    if s in WHITELIST:
        return s, "", ""
    return "总部", s, ""


def copy_style(src_cell, dst_cell):
    """把源单元格样式复制到目标单元格。"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def transform_sheet1(ws):
    """Sheet1: 在 D列(部门) 后插入一列「工区」；按4.1改写 D列。"""
    dept_col = 4  # D
    ws.insert_cols(dept_col + 1)  # 在E处插入新列(原E->F...)
    new_col = dept_col + 1
    # 表头
    ws.cell(row=1, column=new_col, value="工区")
    copy_style(ws.cell(row=1, column=dept_col), ws.cell(row=1, column=new_col))
    # 数据行
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        dept_val = ws.cell(row=r, column=dept_col).value
        gongqu, new_dept = process_sheet1_dept(dept_val)
        # 注意: openpyxl 用 value=None 无法清空已存在值, 必须显式赋空字符串""
        ws.cell(row=r, column=dept_col).value = new_dept
        ws.cell(row=r, column=new_col).value = gongqu
        # 新单元格沿用该行相邻单元格样式
        copy_style(ws.cell(row=r, column=dept_col), ws.cell(row=r, column=new_col))
    # 复制原E列列宽到新列(使其与部门列等宽)
    src_letter = get_column_letter(dept_col)
    new_letter = get_column_letter(new_col)
    if src_letter in ws.column_dimensions:
        ws.column_dimensions[new_letter].width = ws.column_dimensions[src_letter].width


def transform_sheet2(ws):
    """Sheet2: 在 D列(所属单位) 后插入「工区/部门/班组」三列；按4.2填充；原列保留。"""
    unit_col = 4  # D
    ws.insert_cols(unit_col + 1, amount=3)  # 在E/F/G插入三列(原E及以后右移)
    gongqu_col, dept_col_new, banzu_col = unit_col + 1, unit_col + 2, unit_col + 3
    # 第1行是合并大标题 A1:J1, insert_cols 后该 merge 已过时(指向不存在的单元格),
    # 直接重建 merged_cells.ranges 使大标题覆盖全部列
    from openpyxl.utils import range_boundaries, get_column_letter
    from openpyxl.worksheet.cell_range import CellRange
    kept = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        if min_row == 1 and max_row == 1:
            continue  # 丢弃过时的整行大标题, 下面重建
        kept.append(CellRange(str(mr)))
    kept.append(CellRange(f"A1:{get_column_letter(ws.max_column)}1"))
    ws.merged_cells.ranges = kept
    # 表头行 = 第2行(第1行是合并大标题)
    header_row = 2
    headers = ["工区", "部门", "班组"]
    for i, h in enumerate(headers):
        c = ws.cell(row=header_row, column=gongqu_col + i, value=h)
        copy_style(ws.cell(row=header_row, column=unit_col), c)
    # 数据行 3..max_row
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        unit_val = ws.cell(row=r, column=unit_col).value
        gongqu, dept, banzu = process_sheet2_unit(unit_val)
        ws.cell(row=r, column=gongqu_col).value = gongqu
        ws.cell(row=r, column=dept_col_new).value = dept
        ws.cell(row=r, column=banzu_col).value = banzu
        copy_style(ws.cell(row=r, column=unit_col), ws.cell(row=r, column=gongqu_col))
        copy_style(ws.cell(row=r, column=unit_col), ws.cell(row=r, column=dept_col_new))
        copy_style(ws.cell(row=r, column=unit_col), ws.cell(row=r, column=banzu_col))
    # 列宽沿用所属单位列
    src_letter = get_column_letter(unit_col)
    for off in range(3):
        new_letter = get_column_letter(gongqu_col + off)
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[new_letter].width = ws.column_dimensions[src_letter].width


def main():
    wb = openpyxl.load_workbook(SRC)
    ws1 = wb["技能等级 职称"]
    ws2 = wb["考核结果"]
    transform_sheet1(ws1)
    transform_sheet2(ws2)
    wb.save(DST)
    print("已生成:", DST)


if __name__ == "__main__":
    main()
